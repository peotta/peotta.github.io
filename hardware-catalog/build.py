#!/usr/bin/env python3
"""
build.py — converte catalogo.xlsx em assets/data.js para o site estático.

Uso:
    python3 build.py
    python3 build.py minha_planilha.xlsx

Rode este script sempre que atualizar a planilha (novos itens, specs,
fotos ou manuais) e depois faça commit + push do repositório no GitHub.

Colunas opcionais reconhecidas (se você adicionar na planilha, o site
usa automaticamente; se não existirem, o item aparece sem foto/manual):
  - Nas abas de itens (Placas-mãe, Placas de vídeo, Processadores, Memórias):
      "Imagem"       -> nome do arquivo dentro da pasta images/ (ex.: MB-001.jpg)
      "Manual (URL)" -> nome do arquivo dentro da pasta manuals/ (ex.: MB-001.pdf)
                        ou uma URL completa (https://...)
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta a biblioteca openpyxl. Instale com: pip install openpyxl")

XLSX_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("catalogo.xlsx")
OUT_PATH = Path(__file__).parent / "assets" / "data.js"


def sheet_to_records(ws):
    """Converte uma planilha em lista de dicts, usando a primeira linha
    não vazia como cabeçalho (pula título e linhas em branco antes dele)."""
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c not in (None, "")]
        if len(non_empty) >= 3:
            header_idx = i
            break
    if header_idx is None:
        return []
    headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[header_idx])]
    records = []
    for row in rows[header_idx + 1:]:
        if all(c is None for c in row):
            continue
        rec = {}
        for h, v in zip(headers, row):
            if not h or h.startswith("col"):
                continue
            rec[h] = "" if v is None else v
        if rec:
            records.append(rec)
    return records


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"Planilha não encontrada: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    sheets = {name: sheet_to_records(wb[name]) for name in wb.sheetnames}

    data = {
        "motherboards": sheets.get("Placas-mãe", []),
        "gpus": sheets.get("Placas de vídeo", []),
        "cpus": sheets.get("Processadores", []),
        "memory": sheets.get("Memórias", []),
        "tests": sheets.get("Histórico de testes", []),
        "photoIndex": sheets.get("Índice de fotos", []),
        "equipment": sheets.get("Equipamentos", []),
        "guide": sheets.get("Guia", []),
    }

    counts = {
        "motherboards": len(data["motherboards"]),
        "gpus": len(data["gpus"]),
        "cpus": len(data["cpus"]),
        "memory": len(data["memory"]),
    }
    data["meta"] = {"counts": counts}

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    js = "// Gerado automaticamente por build.py — não edite à mão.\n"
    js += "const CATALOG_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n"
    OUT_PATH.write_text(js, encoding="utf-8")

    print(f"OK: {OUT_PATH} gerado com sucesso.")
    for k, v in counts.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
